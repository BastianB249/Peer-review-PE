from datetime import date
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PEERS = [
    ("Aalberts", "AALB.AS", 1, "Closest diversified industrial technology peer"),
    ("Dürr AG", "DUE.DE", 1, "Industrial automation/manufacturing systems exposure comparable to TKH"),
    ("Basler", "BSL.DE", 1, "Direct machine vision hardware/software comparable"),
    ("Cognex", "COGX", 1, "Global machine vision leader benchmark"),
    ("Jenoptik", "JEN.DE", 1, "Photonics and optical systems overlap"),
    ("Huber+Suhner", "HUBN.SW", 1, "Connectivity and industrial cabling exposure"),
    ("NKT", "NKT.CO", 1, "Power/subsea cable and electrification exposure"),
    ("Mersen", "MRN.PA", 1, "Electrical components and power management peer"),
    ("Arcadis", "ARCAD.AS", 0, "Services-heavy model, less product/asset intensity"),
    ("Fugro", "FUR.AS", 0, "Geo-data/offshore services, weaker industrial comparability"),
    ("SBM Offshore", "SBMO.AS", 0, "Project/offshore leasing model not close to TKH"),
    ("Vopak", "VPK.AS", 0, "Tank storage infrastructure, limited operating overlap"),
    ("TKH (subject company)", "TWEKA.AS", 0, "Subject company reference (excluded from peer stats)"),
]


def _cell(col: int, row: int) -> str:
    return f"{get_column_letter(col)}{row}"


def _apply_row_fill(sheet, row: int, max_col: int, fill: PatternFill) -> None:
    for col in range(1, max_col + 1):
        sheet.cell(row=row, column=col).fill = fill


def main() -> None:
    years = [2023, 2024]
    prior_year, latest_year = years
    today = date.today()

    base_columns = [
        "Company",
        "Ticker",
        "Selected (1/0)",
        "Selection rationale",
        "Currency",
        "Share Price (CCY)",
        "Market Cap (CCY m)",
        "Enterprise Value (CCY m)",
        "Net Debt (CCY m)",
        "Equity Beta",
        "FX to EUR",
        "Share Price (EUR)",
        "Market Cap (EUR m)",
        "Enterprise Value (EUR m)",
        "Net Debt (EUR m)",
    ]

    operating_groups = [
        ("Revenue (CCY m)", years),
        ("EBITDA (CCY m)", years),
        ("EBIT (CCY m)", years),
    ]

    multiple_groups = [
        ("EV/Sales", years),
        ("EV/EBITDA", years),
        ("EV/EBIT", years),
    ]

    tail_columns = ["Net Debt/EBITDA"]
    margin_group = ("EBITDA Margin", years)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Peer_Table"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    selected_fill = PatternFill("solid", fgColor="E2F0D9")

    row_1 = 1
    row_2 = 2

    current_col = 1
    base_map: Dict[str, int] = {}
    group_map: Dict[Tuple[str, int], int] = {}

    for label in base_columns:
        sheet.cell(row=row_1, column=current_col, value=label)
        sheet.merge_cells(start_row=row_1, start_column=current_col, end_row=row_2, end_column=current_col)
        base_map[label] = current_col
        current_col += 1

    for group_label, group_years in operating_groups + multiple_groups:
        start_col = current_col
        sheet.cell(row=row_1, column=start_col, value=group_label)
        for year in group_years:
            sheet.cell(row=row_2, column=current_col, value=str(year))
            group_map[(group_label, year)] = current_col
            current_col += 1
        sheet.merge_cells(start_row=row_1, start_column=start_col, end_row=row_1, end_column=current_col - 1)

    for label in tail_columns:
        sheet.cell(row=row_1, column=current_col, value=label)
        sheet.merge_cells(start_row=row_1, start_column=current_col, end_row=row_2, end_column=current_col)
        base_map[label] = current_col
        current_col += 1

    margin_label, margin_years = margin_group
    start_col = current_col
    sheet.cell(row=row_1, column=start_col, value=margin_label)
    for year in margin_years:
        sheet.cell(row=row_2, column=current_col, value=str(year))
        group_map[(margin_label, year)] = current_col
        current_col += 1
    sheet.merge_cells(start_row=row_1, start_column=start_col, end_row=row_1, end_column=current_col - 1)

    last_col = current_col - 1

    for row in range(row_1, row_2 + 1):
        for col in range(1, last_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

    sheet.freeze_panes = "A3"

    data_start_row = 3
    for offset, (name, ticker, selected, rationale) in enumerate(PEERS):
        row = data_start_row + offset
        sheet.cell(row=row, column=base_map["Company"], value=name)
        sheet.cell(row=row, column=base_map["Ticker"], value=ticker)
        sheet.cell(row=row, column=base_map["Selected (1/0)"], value=selected)
        sheet.cell(row=row, column=base_map["Selection rationale"], value=rationale)

        ev_col = base_map["Enterprise Value (CCY m)"]
        net_debt_col = base_map["Net Debt (CCY m)"]

        for year in years:
            revenue_col = group_map[("Revenue (CCY m)", year)]
            ebitda_col = group_map[("EBITDA (CCY m)", year)]
            ebit_col = group_map[("EBIT (CCY m)", year)]
            ev_sales_col = group_map[("EV/Sales", year)]
            ev_ebitda_col = group_map[("EV/EBITDA", year)]
            ev_ebit_col = group_map[("EV/EBIT", year)]
            margin_col = group_map[(margin_label, year)]

            revenue_cell = _cell(revenue_col, row)
            ebitda_cell = _cell(ebitda_col, row)
            ebit_cell = _cell(ebit_col, row)
            ev_cell = _cell(ev_col, row)

            sheet.cell(
                row=row,
                column=ev_sales_col,
                value=f"=IF(OR({revenue_cell}=\"\",{revenue_cell}=0),\"\",{ev_cell}/{revenue_cell})",
            )
            sheet.cell(
                row=row,
                column=ev_ebitda_col,
                value=f"=IF(OR({ebitda_cell}=\"\",{ebitda_cell}=0),\"\",{ev_cell}/{ebitda_cell})",
            )
            sheet.cell(
                row=row,
                column=ev_ebit_col,
                value=f"=IF(OR({ebit_cell}=\"\",{ebit_cell}=0),\"\",{ev_cell}/{ebit_cell})",
            )
            sheet.cell(
                row=row,
                column=margin_col,
                value=f"=IF(OR({revenue_cell}=\"\",{revenue_cell}=0),\"\",{ebitda_cell}/{revenue_cell})",
            )

        latest_ebitda_col = group_map[("EBITDA (CCY m)", latest_year)]
        net_debt_ebitda_col = base_map["Net Debt/EBITDA"]
        latest_ebitda_cell = _cell(latest_ebitda_col, row)
        net_debt_cell = _cell(net_debt_col, row)
        sheet.cell(
            row=row,
            column=net_debt_ebitda_col,
            value=f"=IF(OR({latest_ebitda_cell}=\"\",{latest_ebitda_cell}=0),\"\",{net_debt_cell}/{latest_ebitda_cell})",
        )

        if selected == 1:
            _apply_row_fill(sheet, row, last_col, selected_fill)

    data_end_row = data_start_row + len(PEERS) - 1

    sheet.auto_filter.ref = f"A2:{get_column_letter(last_col)}{data_end_row}"

    summary_start = data_end_row + 2
    avg_row = summary_start
    median_row = summary_start + 1

    sheet.cell(row=avg_row, column=1, value="Selected peers average")
    sheet.cell(row=median_row, column=1, value="Selected peers median")
    sheet.cell(row=avg_row, column=1).font = Font(bold=True)
    sheet.cell(row=median_row, column=1).font = Font(bold=True)

    selected_range = f"{_cell(base_map['Selected (1/0)'], data_start_row)}:{_cell(base_map['Selected (1/0)'], data_end_row)}"

    multiple_columns: List[int] = []
    for label, group_years in multiple_groups:
        for year in group_years:
            multiple_columns.append(group_map[(label, year)])

    for col in multiple_columns:
        col_letter = get_column_letter(col)
        range_ref = f"{col_letter}{data_start_row}:{col_letter}{data_end_row}"
        avg_formula = f"=IFERROR(AVERAGEIF({selected_range},1,{range_ref}),\"\")"
        median_formula = (
            f"=IFERROR(MEDIAN(FILTER({range_ref},({selected_range}=1)*({range_ref}<>\"\"))),\"\")"
        )
        sheet.cell(row=avg_row, column=col, value=avg_formula)
        sheet.cell(row=median_row, column=col, value=median_formula)

    inputs_start = median_row + 3
    sheet.cell(row=inputs_start, column=1, value="TKH Inputs")
    sheet.cell(row=inputs_start, column=1).font = Font(bold=True)

    input_header_row = inputs_start + 1
    sheet.cell(row=input_header_row, column=1, value="Metric")
    sheet.cell(row=input_header_row, column=2, value=str(prior_year))
    sheet.cell(row=input_header_row, column=3, value=str(latest_year))

    input_rows = [
        ("Revenue (CCY m)", "revenue"),
        ("EBITDA (CCY m)", "ebitda"),
        ("EBIT (CCY m)", "ebit"),
        ("Net Debt (CCY m)", "net_debt"),
        ("Adjustments (CCY m)", "adjustments"),
        ("Shares Outstanding (m)", "shares"),
    ]

    input_cells: Dict[str, Dict[str, str]] = {}
    ticker_col = base_map["Ticker"]
    tkh_match = f"MATCH(\"TWEKA.AS\",{_cell(ticker_col, data_start_row)}:{_cell(ticker_col, data_end_row)},0)"
    for idx, (label, key) in enumerate(input_rows, start=1):
        row = input_header_row + idx
        sheet.cell(row=row, column=1, value=label)
        if key in {"revenue", "ebitda", "ebit"}:
            input_cells[key] = {
                "prior": _cell(2, row),
                "latest": _cell(3, row),
            }
            for year_col, year in zip((2, 3), (prior_year, latest_year)):
                source_col = group_map[(label, year)]
                source_range = f"{_cell(source_col, data_start_row)}:{_cell(source_col, data_end_row)}"
                sheet.cell(
                    row=row,
                    column=year_col,
                    value=f"=IFERROR(INDEX({source_range},{tkh_match}),\"\")",
                )
        else:
            cell_ref = _cell(3, row)
            input_cells[key] = {"value": cell_ref}
            if key == "adjustments":
                sheet.cell(row=row, column=3, value=0)
            if key == "net_debt":
                source_col = base_map["Net Debt (CCY m)"]
                source_range = f"{_cell(source_col, data_start_row)}:{_cell(source_col, data_end_row)}"
                sheet.cell(row=row, column=3, value=f"=IFERROR(INDEX({source_range},{tkh_match}),\"\")")

    valuation_start = input_header_row + len(input_rows) + 3
    sheet.cell(row=valuation_start, column=1, value="TKH valuation (Selected peers)")
    sheet.cell(row=valuation_start, column=1).font = Font(bold=True)

    valuation_header_row = valuation_start + 1
    valuation_headers = [
        "Multiple",
        "Year",
        "Selected Avg",
        "Selected Median",
        "TKH Metric (CCY m)",
        "Implied EV (Avg)",
        "Implied EV (Median)",
        "Net Debt (CCY m)",
        "Adjustments (CCY m)",
        "Equity Value (Avg)",
        "Equity Value (Median)",
        "Shares (m)",
        "Per Share (Avg)",
        "Per Share (Median)",
    ]

    for col, header in enumerate(valuation_headers, start=1):
        sheet.cell(row=valuation_header_row, column=col, value=header)

    valuation_rows: List[Tuple[str, str, str]] = []
    for label, metric_key in [
        ("EV/Sales", "revenue"),
        ("EV/EBITDA", "ebitda"),
        ("EV/EBIT", "ebit"),
    ]:
        for year in years:
            valuation_rows.append((label, str(year), metric_key))

    net_debt_ref = input_cells["net_debt"]["value"]
    adjustments_ref = input_cells["adjustments"]["value"]
    shares_ref = input_cells["shares"]["value"]

    for idx, (multiple_label, year_label, metric_key) in enumerate(valuation_rows, start=1):
        row = valuation_header_row + idx
        sheet.cell(row=row, column=1, value=multiple_label)
        sheet.cell(row=row, column=2, value=year_label)

        year = int(year_label)
        avg_multiple_cell = _cell(group_map[(multiple_label, year)], avg_row)
        median_multiple_cell = _cell(group_map[(multiple_label, year)], median_row)
        metric_cell = input_cells[metric_key]["prior" if year == prior_year else "latest"]

        sheet.cell(row=row, column=3, value=f"={avg_multiple_cell}")
        sheet.cell(row=row, column=4, value=f"={median_multiple_cell}")
        sheet.cell(row=row, column=5, value=f"={metric_cell}")

        avg_implied_ev = _cell(6, row)
        median_implied_ev = _cell(7, row)
        avg_equity = _cell(10, row)
        median_equity = _cell(11, row)

        sheet.cell(
            row=row,
            column=6,
            value=f"=IF(OR(C{row}=\"\",E{row}=\"\"),\"\",C{row}*E{row})",
        )
        sheet.cell(
            row=row,
            column=7,
            value=f"=IF(OR(D{row}=\"\",E{row}=\"\"),\"\",D{row}*E{row})",
        )
        sheet.cell(row=row, column=8, value=f"={net_debt_ref}")
        sheet.cell(row=row, column=9, value=f"={adjustments_ref}")
        sheet.cell(
            row=row,
            column=10,
            value=f"=IF({avg_implied_ev}=\"\",\"\",{avg_implied_ev}-H{row}+I{row})",
        )
        sheet.cell(
            row=row,
            column=11,
            value=f"=IF({median_implied_ev}=\"\",\"\",{median_implied_ev}-H{row}+I{row})",
        )
        sheet.cell(row=row, column=12, value=f"={shares_ref}")
        sheet.cell(
            row=row,
            column=13,
            value=f"=IF(OR({avg_equity}=\"\",L{row}=\"\"),\"\",{avg_equity}/L{row})",
        )
        sheet.cell(
            row=row,
            column=14,
            value=f"=IF(OR({median_equity}=\"\",L{row}=\"\"),\"\",{median_equity}/L{row})",
        )

    column_widths = {
        1: 18,
        2: 12,
        3: 14,
        4: 46,
        5: 10,
        6: 14,
        7: 18,
        8: 20,
        9: 18,
        10: 10,
        11: 18,
        12: 20,
        13: 22,
        14: 18,
    }
    for col, width in column_widths.items():
        sheet.column_dimensions[get_column_letter(col)].width = width

    instructions = workbook.create_sheet(title="Instructions")
    instructions_data = [
        ["TKH peer workbook", f"Generated on {today.isoformat()}"],
        [
            "How to populate live KPIs",
            "Run fill_from_yahoo.py or paste current values for market data and operating metrics.",
        ],
        ["Suggested data sources", "Bloomberg, Capital IQ, FactSet, Refinitiv, or Yahoo Finance."],
        [
            "Units",
            "Market Cap/EV/Revenue/EBITDA/EBIT/Net Debt are stored in CCY m; shares in m; EUR columns use FX.",
        ],
        ["Selection result", "Rows with Selected=1 drive the peer summary and valuation block."],
    ]
    for row, values in enumerate(instructions_data, start=1):
        for col, value in enumerate(values, start=1):
            instructions.cell(row=row, column=col, value=value)

    wacc = workbook.create_sheet(title="WACC_Model")
    wacc_headers = [
        "Company",
        "Ticker",
        "Selected (1/0)",
        "Equity Beta",
        "Net Debt (CCY m)",
        "Market Cap (CCY m)",
        "Debt/Equity",
        "Unlevered Beta",
    ]
    for col, header in enumerate(wacc_headers, start=1):
        wacc.cell(row=1, column=col, value=header).font = Font(bold=True)

    for idx, _ in enumerate(PEERS, start=0):
        peer_row = data_start_row + idx
        wacc_row = 2 + idx
        for col, header in enumerate(wacc_headers, start=1):
            if header in base_map:
                source_col = base_map[header]
                wacc.cell(row=wacc_row, column=col, value=f"=Peer_Table!{_cell(source_col, peer_row)}")
        wacc.cell(row=wacc_row, column=1, value=f"=Peer_Table!{_cell(base_map['Company'], peer_row)}")
        wacc.cell(row=wacc_row, column=2, value=f"=Peer_Table!{_cell(base_map['Ticker'], peer_row)}")
        wacc.cell(row=wacc_row, column=3, value=f"=Peer_Table!{_cell(base_map['Selected (1/0)'], peer_row)}")
        wacc.cell(row=wacc_row, column=4, value=f"=Peer_Table!{_cell(base_map['Equity Beta'], peer_row)}")
        wacc.cell(row=wacc_row, column=5, value=f"=Peer_Table!{_cell(base_map['Net Debt (CCY m)'], peer_row)}")
        wacc.cell(row=wacc_row, column=6, value=f"=Peer_Table!{_cell(base_map['Market Cap (CCY m)'], peer_row)}")
        wacc.cell(
            row=wacc_row,
            column=7,
            value=f"=IF(OR(F{wacc_row}=\"\",F{wacc_row}=0,E{wacc_row}=\"\"),\"\",E{wacc_row}/F{wacc_row})",
        )

    input_start = len(PEERS) + 4
    wacc.cell(row=input_start, column=1, value="WACC Inputs").font = Font(bold=True)
    inputs = [
        ("Risk-free rate", 0.03),
        ("Market risk premium", 0.045),
        ("Small firm premium", 0.0125),
        ("Marginal tax rate", 0.25),
        ("Cost of debt (pre-tax)", 0.05),
        ("Target debt / equity", 0.25),
    ]
    for idx, (label, default) in enumerate(inputs, start=1):
        row = input_start + idx
        wacc.cell(row=row, column=1, value=label)
        wacc.cell(row=row, column=2, value=default)

    summary_start = input_start + len(inputs) + 2
    wacc.cell(row=summary_start, column=1, value="Selected peers beta summary").font = Font(bold=True)
    wacc.cell(row=summary_start + 1, column=1, value="Average equity beta")
    wacc.cell(row=summary_start + 2, column=1, value="Median equity beta")
    wacc.cell(row=summary_start + 3, column=1, value="Average unlevered beta")
    wacc.cell(row=summary_start + 4, column=1, value="Median unlevered beta")

    selected_range_wacc = f"C2:C{len(PEERS)+1}"
    equity_beta_range = f"D2:D{len(PEERS)+1}"
    unlevered_start_row = 2
    tax_rate_cell = f"B{input_start + 4}"
    for idx in range(len(PEERS)):
        row = unlevered_start_row + idx
        wacc.cell(
            row=row,
            column=8,
            value=f"=IF(OR(D{row}=\"\",G{row}=\"\"),\"\",D{row}/(1+(1-{tax_rate_cell})*G{row}))",
        )

    unlevered_range = f"H2:H{len(PEERS)+1}"
    wacc.cell(
        row=summary_start + 1,
        column=2,
        value=f"=IFERROR(AVERAGEIF({selected_range_wacc},1,{equity_beta_range}),\"\")",
    )
    wacc.cell(
        row=summary_start + 2,
        column=2,
        value=f"=IFERROR(MEDIAN(FILTER({equity_beta_range},({selected_range_wacc}=1)*({equity_beta_range}<>\"\"))),\"\")",
    )
    wacc.cell(
        row=summary_start + 3,
        column=2,
        value=f"=IFERROR(AVERAGEIF({selected_range_wacc},1,{unlevered_range}),\"\")",
    )
    wacc.cell(
        row=summary_start + 4,
        column=2,
        value=f"=IFERROR(MEDIAN(FILTER({unlevered_range},({selected_range_wacc}=1)*({unlevered_range}<>\"\"))),\"\")",
    )

    calc_start = summary_start + 6
    wacc.cell(row=calc_start, column=1, value="WACC Calculation").font = Font(bold=True)
    calc_rows = [
        ("Relevered beta", f"=B{summary_start + 3}*(1+(1-B{input_start + 4})*B{input_start + 6})"),
        ("Cost of equity", f"=B{input_start + 1}+B{input_start + 2}*B{calc_start + 1}+B{input_start + 3}"),
        ("Cost of debt (after-tax)", f"=B{input_start + 5}*(1-B{input_start + 4})"),
        ("Target debt weight", f"=B{input_start + 6}/(1+B{input_start + 6})"),
        ("Target equity weight", f"=1-B{calc_start + 4}"),
        (
            "WACC",
            f"=B{calc_start + 2}*B{calc_start + 5}+B{calc_start + 3}*B{calc_start + 4}",
        ),
    ]
    for idx, (label, formula) in enumerate(calc_rows, start=1):
        row = calc_start + idx
        wacc.cell(row=row, column=1, value=label)
        wacc.cell(row=row, column=2, value=formula)

    wacc.column_dimensions["A"].width = 28
    wacc.column_dimensions["B"].width = 18
    for col in range(3, 9):
        wacc.column_dimensions[get_column_letter(col)].width = 18

    workbook.save("TKH_Peer_Analysis.xlsx")


if __name__ == "__main__":
    main()
