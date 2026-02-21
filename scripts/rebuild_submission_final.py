from __future__ import annotations

from statistics import mean, median
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

WRDS_FILE = Path('outputs/TKH_Peer_Analysis_submission_ready.xlsx')
YAHOO_FILE = Path('TKH_Peer_Analysis_filled.xlsx')
OUT_FILE = Path('outputs/TKH_Peer_Analysis_submission_ready_FINAL.xlsx')

HEADER = PatternFill('solid', fgColor='0B2A4A')
SUBHEADER = PatternFill('solid', fgColor='D9D9D9')
WHITE = Font(color='FFFFFF', bold=True)
BOLD = Font(bold=True)


def load_rows() -> list[dict]:
    wrds = load_workbook(WRDS_FILE, data_only=True)['Peer_Table']
    yahoo = load_workbook(YAHOO_FILE, data_only=True)['Peer_Table']

    rows: list[dict] = []
    for r in range(2, 11):
        rows.append(
            {
                'company': wrds.cell(r, 1).value,
                'ticker': wrds.cell(r, 2).value,
                'selected': int(wrds.cell(r, 3).value or 0),
                'rationale': wrds.cell(r, 7).value,
                'currency': wrds.cell(r, 8).value,
                'fx': float(wrds.cell(r, 16).value or 1),
                'price': wrds.cell(r, 9).value,
                'mcap': wrds.cell(r, 10).value,
                'ev': wrds.cell(r, 11).value,
                'net_debt': wrds.cell(r, 14).value,
                'beta': wrds.cell(r, 15).value,
                'rev_2023': wrds.cell(r, 17).value,
                'ebitda_2023': wrds.cell(r, 18).value,
                'ebit_2023': wrds.cell(r, 19).value,
                'rev_2024': wrds.cell(r, 23).value,
                'ebitda_2024': wrds.cell(r, 24).value,
                'ebit_2024': wrds.cell(r, 25).value,
                'source': 'WRDS',
            }
        )

    # Replace Cognex with Yahoo values from previous final draft as requested.
    cgx = next(row for row in rows if row['company'] == 'Cognex')
    for r in range(3, 20):
        if yahoo.cell(r, 1).value == 'Cognex':
            cgx['ticker'] = 'CGNX'
            cgx['currency'] = yahoo.cell(r, 5).value
            cgx['price'] = yahoo.cell(r, 6).value
            cgx['mcap'] = yahoo.cell(r, 7).value
            cgx['ev'] = yahoo.cell(r, 8).value
            cgx['net_debt'] = yahoo.cell(r, 9).value
            cgx['beta'] = yahoo.cell(r, 10).value
            cgx['fx'] = yahoo.cell(r, 11).value
            cgx['rev_2023'] = yahoo.cell(r, 16).value
            cgx['rev_2024'] = yahoo.cell(r, 17).value
            cgx['ebitda_2023'] = yahoo.cell(r, 18).value
            cgx['ebitda_2024'] = yahoo.cell(r, 19).value
            cgx['ebit_2023'] = yahoo.cell(r, 20).value
            cgx['ebit_2024'] = yahoo.cell(r, 21).value
            cgx['source'] = 'Yahoo (prior final poll)'
            break

    # Keep only peers + subject in final tabs (no extra names).
    return rows


def style_header(ws, row: int, start_col: int, end_col: int, title: str | None = None) -> None:
    if title is not None:
        ws.cell(row, start_col, title)
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row, c)
        cell.fill = HEADER
        cell.font = WHITE
        cell.alignment = Alignment(horizontal='center')


def build_wacc_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.active
    ws.title = 'WACC_Model'

    tax = 0.225
    target_de = 0.25
    rf = 0.03
    credit_spread = 0.025
    cost_debt = rf + credit_spread
    mkt_rp = 0.045
    sfp = 0.0125

    peers = [r for r in rows if r['selected'] == 1 and 'subject' not in r['company'].lower()]

    ws['A1'] = 'Weighted Average Cost of Capital'
    ws['A1'].font = WHITE
    ws['A1'].fill = HEADER
    ws['D1'] = 'Corporate'
    ws['D1'].font = WHITE
    ws['D1'].fill = HEADER

    labels = [
        ('Riskfree rate', rf),
        ('Risk premium', credit_spread),
        ('Cost of debt', cost_debt),
        ('Marginal tax rate', tax),
        ('Cost of debt after taxes', cost_debt * (1 - tax)),
        (None, None),
        ('Riskfree rate', rf),
        ('Market risk premium', mkt_rp),
        (None, None),
    ]
    row_ptr = 3
    for label, value in labels:
        if label:
            ws[f'A{row_ptr}'] = label
            ws[f'D{row_ptr}'] = value
            ws[f'D{row_ptr}'].number_format = '0.0%'
            if 'Cost of debt' in label:
                ws[f'A{row_ptr}'].font = BOLD
        row_ptr += 1

    # Peer table
    style_header(ws, 1, 6, 11, 'PEER GROUP')
    ws['I1'] = 'Equity beta'
    ws['J1'] = 'Av. D/E'
    ws['K1'] = 'Unlev. Beta'

    unlev_betas = []
    lev_betas = []
    des = []
    out = 3
    for p in peers:
        de = None
        if p['mcap'] not in (None, 0) and p['net_debt'] is not None:
            de = p['net_debt'] / p['mcap']
        ub = None
        if de is not None and p['beta'] is not None:
            ub = p['beta'] / (1 + (1 - tax) * de)
        ws[f'F{out}'] = p['company']
        ws[f'I{out}'] = p['beta']
        ws[f'J{out}'] = de
        ws[f'K{out}'] = ub
        if p['beta'] is not None:
            lev_betas.append(p['beta'])
        if de is not None:
            des.append(de)
        if ub is not None:
            unlev_betas.append(ub)
        out += 1

    ws[f'F{out+1}'] = 'Average'
    ws[f'F{out+2}'] = 'Median'
    ws[f'F{out+1}'].font = BOLD
    ws[f'F{out+2}'].font = BOLD
    ws[f'I{out+1}'] = mean(lev_betas)
    ws[f'I{out+2}'] = median(lev_betas)
    ws[f'J{out+1}'] = mean(des)
    ws[f'J{out+2}'] = median(des)
    ws[f'K{out+1}'] = mean(unlev_betas)
    ws[f'K{out+2}'] = median(unlev_betas)

    unlev = median(unlev_betas)
    relevered = unlev * (1 + (1 - tax) * target_de)
    cost_equity = rf + relevered * mkt_rp + sfp
    wacc = cost_equity * (1 / (1 + target_de)) + (cost_debt * (1 - tax)) * (target_de / (1 + target_de))

    ws['A12'] = 'Unlevered beta'
    ws['D12'] = unlev
    ws['A13'] = 'Target D/E'
    ws['D13'] = target_de
    ws['A14'] = 'Relevered beta'
    ws['D14'] = relevered
    ws['A14'].font = BOLD
    ws['A15'] = 'Small firm premium'
    ws['D15'] = sfp
    ws['A16'] = 'Cost of common equity'
    ws['D16'] = cost_equity
    ws['A16'].font = BOLD
    ws['A18'] = 'Cost of preferred equity'
    ws['D18'] = 0
    ws['A20'] = 'Target interestbearing debt'
    ws['D20'] = target_de / (1 + target_de)
    ws['A21'] = 'Target preferred equity'
    ws['D21'] = 0
    ws['A22'] = 'Target common equity'
    ws['D22'] = 1 - ws['D20'].value
    ws['A24'] = 'WACC'
    ws['D24'] = wacc
    ws['A24'].font = BOLD
    ws['D24'].font = BOLD

    for col in ['D', 'I', 'J', 'K']:
        for r in range(3, 25):
            if ws[f'{col}{r}'].value is not None and isinstance(ws[f'{col}{r}'].value, (float, int)):
                ws[f'{col}{r}'].number_format = '0.0%'


def build_cca_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet('CCA_Model')
    style_header(ws, 1, 1, 12, 'MULTIPLE ANALYSIS')
    headers = ['Company', 'Stock price', 'Market cap', 'Ent. Value', 'EV/Sales 2023', 'EV/Sales 2024', 'EV/EBITDA 2023', 'EV/EBITDA 2024', 'EV/EBIT 2023', 'EV/EBIT 2024', 'Source', 'Selected']
    for c, h in enumerate(headers, 1):
        ws.cell(2, c, h)
        ws.cell(2, c).fill = HEADER
        ws.cell(2, c).font = WHITE

    peers = [r for r in rows if 'subject' not in r['company'].lower()]
    r0 = 3
    for i, p in enumerate(peers):
        ev = p['ev']
        s23 = ev / p['rev_2023'] if p['rev_2023'] else None
        s24 = ev / p['rev_2024'] if p['rev_2024'] else None
        e23 = ev / p['ebitda_2023'] if p['ebitda_2023'] else None
        e24 = ev / p['ebitda_2024'] if p['ebitda_2024'] else None
        b23 = ev / p['ebit_2023'] if p['ebit_2023'] else None
        b24 = ev / p['ebit_2024'] if p['ebit_2024'] else None
        vals = [p['company'], p['price'], p['mcap'], ev, s23, s24, e23, e24, b23, b24, p['source'], p['selected']]
        for c, v in enumerate(vals, 1):
            ws.cell(r0 + i, c, v)

    # Average and median on selected peers only.
    selected = [p for p in peers if p['selected'] == 1]
    avg_r = r0 + len(peers) + 1
    med_r = avg_r + 1
    ws[f'A{avg_r}'] = 'Average'
    ws[f'A{med_r}'] = 'Median'
    ws[f'A{avg_r}'].font = BOLD
    ws[f'A{med_r}'].font = BOLD

    for col, key in [(5, 'rev_2023'), (6, 'rev_2024'), (7, 'ebitda_2023'), (8, 'ebitda_2024'), (9, 'ebit_2023'), (10, 'ebit_2024')]:
        vals = []
        for p in selected:
            denom = p[key]
            vals.append(p['ev'] / denom if denom else None)
        vals = [v for v in vals if v is not None]
        ws.cell(avg_r, col, mean(vals))
        ws.cell(med_r, col, median(vals))

    for r in range(3, med_r + 1):
        for c in [5, 6, 7, 8, 9, 10]:
            if isinstance(ws.cell(r, c).value, (float, int)):
                ws.cell(r, c).number_format = '0.0x'


def build_rationale_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet('Peer_Rationale')
    style_header(ws, 1, 1, 16, 'Peer rationale + raw data')
    headers = [
        'Company', 'Ticker', 'Selected', 'Rationale', 'Currency', 'FX', 'Price', 'MCap', 'EV', 'NetDebt',
        'Beta', 'Revenue 2023', 'EBITDA 2023', 'EBIT 2023', 'Revenue 2024', 'Source'
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(2, c, h)
        ws.cell(2, c).fill = SUBHEADER
        ws.cell(2, c).font = BOLD

    r0 = 3
    for i, p in enumerate(rows):
        vals = [
            p['company'], p['ticker'], p['selected'], p['rationale'], p['currency'], p['fx'], p['price'], p['mcap'], p['ev'], p['net_debt'],
            p['beta'], p['rev_2023'], p['ebitda_2023'], p['ebit_2023'], p['rev_2024'], p['source']
        ]
        for c, v in enumerate(vals, 1):
            ws.cell(r0 + i, c, v)


def main() -> None:
    rows = load_rows()
    wb = Workbook()
    build_wacc_sheet(wb, rows)
    build_cca_sheet(wb, rows)
    build_rationale_sheet(wb, rows)
    wb.save(OUT_FILE)
    print(f'Wrote {OUT_FILE}')


if __name__ == '__main__':
    main()
