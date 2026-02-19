from __future__ import annotations

import csv
import logging
import math
import os
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ------------------------------
# Configurable assumptions
# ------------------------------
AS_OF_OVERRIDE = None
FISCAL_YEARS = [2023, 2024]
OUTPUT_FILE = Path("outputs/TKH_Peer_Analysis_submission_ready.xlsx")
PEER_INPUT_FILE = Path("inputs/peer_universe.csv")
WRDS_MAPPING_FILE = Path("inputs/wrds_mapping.csv")
OVERRIDE_FILE = Path("inputs/data_overrides.csv")

USE_PROVIDER_EV_AS_TRUTH = True
ALLOW_MIXED_SOURCES = True  # if WRDS misses a field, allow Yahoo fallback for that field

RISK_FREE_RATE = 0.030
EQUITY_RISK_PREMIUM = 0.050
SMALL_FIRM_PREMIUM = 0.0125
MARGINAL_TAX_RATE = 0.25
COST_OF_DEBT_PRE_TAX = 0.055
TARGET_D_OVER_E = 0.25

BETA_HORIZON = "5Y"
BETA_FREQUENCY = "Monthly"
BETA_INDEX = "MSCI World"
COST_OF_DEBT_METHOD = "Assumption (Rf + spread proxy); replace with issuer-implied yield when available"
ERP_SOURCE_NOTE = "KPMG cost of capital study (manual input required in this environment)"
SFP_SOURCE_NOTE = "KPMG size premium table (manual input required in this environment)"

LOG_FILE = Path("outputs/build_peer_model.log")


@dataclass
class WrdsMapping:
    ticker: str
    region: str
    wrds_db: str
    identifier_type: str
    identifier_value: str
    notes: str = ""


@dataclass
class WrdsPullStatus:
    connected: bool = False
    connection_message: str = ""
    mapping_coverage: int = 0
    per_peer_message: dict[str, str] = field(default_factory=dict)
    statement_filters: str = (
        "Compustat annual; filters: indfmt='INDL', datafmt='STD', consol='C' "
        "(+ popsrc='D' for North America when available)."
    )


@dataclass
class PeerRow:
    company: str
    ticker: str
    selected: int
    core_set: int
    segment_fit: str
    peer_status: str
    selection_rationale: str
    gvkey: str = ""
    currency: str | None = None
    share_price_ccy: float | None = None
    market_cap_ccy_m: float | None = None
    enterprise_value_ccy_m: float | None = None
    gross_debt_ccy_m: float | None = None
    cash_ccy_m: float | None = None
    net_debt_ccy_m: float | None = None
    equity_beta: float | None = None
    fx_to_eur: float | None = None
    revenue: dict[int, float | None] = field(default_factory=dict)
    ebitda: dict[int, float | None] = field(default_factory=dict)
    ebit: dict[int, float | None] = field(default_factory=dict)
    source_by_field: dict[str, str] = field(default_factory=dict)

    @property
    def enterprise_value_input_ccy_m(self) -> float | None:
        if USE_PROVIDER_EV_AS_TRUTH:
            return self.enterprise_value_ccy_m
        if self.market_cap_ccy_m is None or self.net_debt_ccy_m is None:
            return None
        return self.market_cap_ccy_m + self.net_debt_ccy_m


def setup_logging() -> None:
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[logging.FileHandler(LOG_FILE, mode="w"), logging.StreamHandler()],
    )


def to_m(v: Any) -> float | None:
    if v in (None, ""):
        return None
    try:
        return float(v) / 1_000_000
    except Exception:
        return None


def parse_peers(path: Path) -> list[PeerRow]:
    peers: list[PeerRow] = []
    with path.open("r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            peers.append(
                PeerRow(
                    company=row["company"],
                    ticker=row["ticker"],
                    selected=int(row["selected"]),
                    core_set=int(row["core_set"]),
                    segment_fit=row["segment_fit"],
                    peer_status=row["peer_status"],
                    selection_rationale=row["selection_rationale"],
                    gvkey=row.get("gvkey", "") or "",
                    revenue={y: None for y in FISCAL_YEARS},
                    ebitda={y: None for y in FISCAL_YEARS},
                    ebit={y: None for y in FISCAL_YEARS},
                )
            )
    return peers


def parse_wrds_mapping(path: Path) -> dict[str, WrdsMapping]:
    if not path.exists():
        return {}
    df = pd.read_csv(path, dtype=str).fillna("")
    required = {"ticker", "region", "wrds_db", "identifier_type", "identifier_value", "notes"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Missing columns in {path}: {sorted(missing)}")
    out: dict[str, WrdsMapping] = {}
    for _, r in df.iterrows():
        out[r["ticker"].strip()] = WrdsMapping(
            ticker=r["ticker"].strip(),
            region=r["region"].strip(),
            wrds_db=r["wrds_db"].strip(),
            identifier_type=r["identifier_type"].strip().lower(),
            identifier_value=r["identifier_value"].strip(),
            notes=r["notes"].strip(),
        )
    return out


def _find_row_label(index: list[Any], labels: list[str]) -> Any | None:
    mapping = {str(lbl).lower(): lbl for lbl in index}
    for label in labels:
        found = mapping.get(label.lower())
        if found is not None:
            return found
    return None


def _extract_metric_by_year(financials: pd.DataFrame, labels: list[str]) -> dict[int, float]:
    if financials is None or financials.empty:
        return {}
    row_lbl = _find_row_label(list(financials.index), labels)
    if row_lbl is None:
        return {}
    out: dict[int, float] = {}
    for col, val in financials.loc[row_lbl].items():
        if pd.isna(val):
            continue
        try:
            out[pd.Timestamp(col).year] = float(val)
        except Exception:
            continue
    return out


def _extract_latest_balance(balance: pd.DataFrame, labels: list[str]) -> float | None:
    if balance is None or balance.empty:
        return None
    row_lbl = _find_row_label(list(balance.index), labels)
    if row_lbl is None:
        return None
    col = balance.columns[0]
    val = balance.loc[row_lbl].get(col)
    if val is None or pd.isna(val):
        return None
    return float(val)


def _last_close(tkr: yf.Ticker) -> float | None:
    try:
        hist = tkr.history(period="5d")
        if hist.empty:
            return None
        return float(hist["Close"].dropna().iloc[-1])
    except Exception:
        return None


def _fetch_fx_rate(ccy: str | None, cache: dict[str, float]) -> float | None:
    if ccy in (None, ""):
        return None
    if ccy == "EUR":
        return 1.0
    if ccy in cache:
        return cache[ccy]
    direct = _last_close(yf.Ticker(f"{ccy}EUR=X"))
    if direct is not None:
        cache[ccy] = direct
        return direct
    inverse = _last_close(yf.Ticker(f"EUR{ccy}=X"))
    if inverse not in (None, 0):
        cache[ccy] = 1.0 / float(inverse)
        return cache[ccy]
    return None


def _set_if_missing(peer: PeerRow, attr: str, value: Any, source: str) -> None:
    current = getattr(peer, attr)
    if current is None and value is not None:
        setattr(peer, attr, value)
        peer.source_by_field[attr] = source


def _set_metric_if_missing(peer: PeerRow, metric: str, year: int, value: Any, source: str) -> None:
    if getattr(peer, metric)[year] is None and value is not None:
        getattr(peer, metric)[year] = value
        peer.source_by_field[f"{metric}_{year}"] = source


def fetch_from_wrds(peers: list[PeerRow], wrds_mapping: dict[str, WrdsMapping]) -> WrdsPullStatus:
    status = WrdsPullStatus()
    username = os.getenv("WRDS_USERNAME")
    if not username:
        status.connection_message = "WRDS skipped: WRDS_USERNAME not set"
        for p in peers:
            status.per_peer_message[p.ticker] = "WRDS skipped (no WRDS_USERNAME)"
        return status

    mapped = [p for p in peers if p.ticker in wrds_mapping and wrds_mapping[p.ticker].identifier_value]
    status.mapping_coverage = len(mapped)

    try:
        import wrds  # type: ignore

        db = wrds.Connection(wrds_username=username)
        status.connected = True
        status.connection_message = "Connected"
    except Exception as exc:
        status.connection_message = f"Connection failed: {exc}"
        for p in peers:
            status.per_peer_message[p.ticker] = "WRDS unavailable; fallback to Yahoo"
        return status

    for p in peers:
        m = wrds_mapping.get(p.ticker)
        if m is None:
            status.per_peer_message[p.ticker] = "No mapping row; fallback to Yahoo"
            continue
        if not m.identifier_value:
            status.per_peer_message[p.ticker] = "Mapping incomplete (identifier_value blank); fallback to Yahoo"
            continue
        if m.identifier_type != "gvkey":
            status.per_peer_message[p.ticker] = f"Unsupported identifier_type={m.identifier_type}; require gvkey"
            continue

        if m.wrds_db == "comp_na":
            schema, table = "comp", "funda"
            where_extra = "and indfmt='INDL' and datafmt='STD' and consol='C' and popsrc='D'"
        elif m.wrds_db == "comp_global":
            schema, table = "compg", "g_funda"
            where_extra = "and indfmt='INDL' and datafmt='STD' and consol='C'"
        else:
            status.per_peer_message[p.ticker] = f"Unsupported wrds_db={m.wrds_db}"
            continue

        sql = f"""
            select gvkey, fyear, datadate, curcd,
                   coalesce(sale, revt) as revenue,
                   coalesce(ebitda, oibdp) as ebitda,
                   coalesce(ebit, oiadp) as ebit,
                   dltt, dlc, che
            from {schema}.{table}
            where gvkey = %(gvkey)s
              and fyear in %(years)s
              {where_extra}
            order by fyear, datadate desc
        """
        try:
            df = db.raw_sql(sql, params={"gvkey": m.identifier_value, "years": tuple(FISCAL_YEARS)})
        except Exception as exc:
            status.per_peer_message[p.ticker] = f"WRDS query failed: {exc}"
            continue

        if df.empty:
            status.per_peer_message[p.ticker] = "WRDS query returned no rows"
            continue

        # deterministic selection for duplicates per fiscal year: latest datadate row
        df = (
            df.sort_values(["fyear", "datadate"], ascending=[True, False])
            .drop_duplicates(subset=["fyear"], keep="first")
            .reset_index(drop=True)
        )

        used_years: list[str] = []
        for _, r in df.iterrows():
            y = int(r["fyear"])
            if y not in FISCAL_YEARS:
                continue
            used_years.append(str(y))
            _set_metric_if_missing(p, "revenue", y, to_m(r.get("revenue")), f"WRDS {schema}.{table}")
            _set_metric_if_missing(p, "ebitda", y, to_m(r.get("ebitda")), f"WRDS {schema}.{table}")
            _set_metric_if_missing(p, "ebit", y, to_m(r.get("ebit")), f"WRDS {schema}.{table}")
            _set_if_missing(p, "currency", r.get("curcd"), f"WRDS {schema}.{table}")

            dltt = to_m(r.get("dltt")) or 0.0
            dlc = to_m(r.get("dlc")) or 0.0
            che = to_m(r.get("che"))
            _set_if_missing(p, "gross_debt_ccy_m", dltt + dlc, f"WRDS {schema}.{table}")
            _set_if_missing(p, "cash_ccy_m", che, f"WRDS {schema}.{table}")
            if (dltt + dlc) is not None and che is not None:
                _set_if_missing(p, "net_debt_ccy_m", (dltt + dlc) - che, f"WRDS {schema}.{table}")

        status.per_peer_message[p.ticker] = f"WRDS success ({schema}.{table}, years={','.join(used_years) or 'none'})"
        logging.info("%s: WRDS source used for fundamentals/debt (%s.%s), chosen latest datadate per fiscal year", p.ticker, schema, table)

    db.close()
    return status


def fetch_from_yahoo(peers: list[PeerRow], wrds_status: WrdsPullStatus) -> None:
    logging.info("Fetching Yahoo data for market fields and fallback fields...")
    fx_cache: dict[str, float] = {}

    for p in peers:
        try:
            tkr = yf.Ticker(p.ticker)
            info = tkr.get_info() or {}
            fin = tkr.financials
            bal = tkr.balance_sheet
        except Exception as exc:
            logging.warning("%s: Yahoo fetch failed: %s", p.ticker, exc)
            wrds_msg = wrds_status.per_peer_message.get(p.ticker, "")
            if wrds_msg.startswith("WRDS success"):
                logging.info("%s: WRDS used; Yahoo supplement unavailable due fetch error", p.ticker)
            else:
                logging.info("%s: No WRDS data and Yahoo unavailable (will remain missing)", p.ticker)
            continue

        # Always Yahoo for these market fields in current model
        if p.share_price_ccy is None and _last_close(tkr) is not None:
            p.share_price_ccy = _last_close(tkr)
            p.source_by_field["share_price_ccy"] = f"Yahoo Finance ({p.ticker})"
        if p.market_cap_ccy_m is None and to_m(info.get("marketCap")) is not None:
            p.market_cap_ccy_m = to_m(info.get("marketCap"))
            p.source_by_field["market_cap_ccy_m"] = f"Yahoo Finance ({p.ticker})"
        if p.enterprise_value_ccy_m is None and to_m(info.get("enterpriseValue")) is not None:
            p.enterprise_value_ccy_m = to_m(info.get("enterpriseValue"))
            p.source_by_field["enterprise_value_ccy_m"] = f"Yahoo Finance ({p.ticker})"
        if p.equity_beta is None and info.get("beta") is not None:
            p.equity_beta = info.get("beta")
            p.source_by_field["equity_beta"] = f"Yahoo Finance ({p.ticker})"

        if p.currency is None and info.get("currency") is not None:
            p.currency = info.get("currency")
            p.source_by_field["currency"] = f"Yahoo Finance ({p.ticker})"

        # fallback financials from Yahoo only if allowed
        rev = _extract_metric_by_year(fin, ["Total Revenue", "TotalRevenue"])
        ebitda = _extract_metric_by_year(fin, ["EBITDA"])
        ebit = _extract_metric_by_year(fin, ["EBIT", "Operating Income", "OperatingIncome"])
        if not ebitda:
            da = _extract_metric_by_year(fin, ["Depreciation And Amortization", "Depreciation & Amortization"])
            for y in set(ebit) & set(da):
                ebitda[y] = ebit[y] + da[y]

        if ALLOW_MIXED_SOURCES:
            for y in FISCAL_YEARS:
                _set_metric_if_missing(p, "revenue", y, to_m(rev.get(y)), f"Yahoo Finance ({p.ticker})")
                _set_metric_if_missing(p, "ebitda", y, to_m(ebitda.get(y)), f"Yahoo Finance ({p.ticker})")
                _set_metric_if_missing(p, "ebit", y, to_m(ebit.get(y)), f"Yahoo Finance ({p.ticker})")

            total_debt = _extract_latest_balance(bal, ["Total Debt", "Long Term Debt", "Current Debt"])
            cash = _extract_latest_balance(
                bal,
                ["Cash And Cash Equivalents", "Cash And Cash Equivalents Including Short Term Investments", "Cash"],
            )
            _set_if_missing(p, "gross_debt_ccy_m", to_m(total_debt), f"Yahoo Finance ({p.ticker})")
            _set_if_missing(p, "cash_ccy_m", to_m(cash), f"Yahoo Finance ({p.ticker})")
            net_debt_info = to_m(info.get("netDebt"))
            if net_debt_info is not None:
                _set_if_missing(p, "net_debt_ccy_m", net_debt_info, f"Yahoo Finance ({p.ticker})")
            elif p.gross_debt_ccy_m is not None and p.cash_ccy_m is not None:
                _set_if_missing(p, "net_debt_ccy_m", p.gross_debt_ccy_m - p.cash_ccy_m, f"Yahoo Finance ({p.ticker})")

        if p.fx_to_eur is None:
            fx = _fetch_fx_rate(p.currency, fx_cache)
            if fx is not None:
                p.fx_to_eur = fx
                p.source_by_field["fx_to_eur"] = f"Yahoo FX ({p.currency}EUR)"

        wrds_msg = wrds_status.per_peer_message.get(p.ticker, "")
        if wrds_msg.startswith("WRDS success"):
            logging.info("%s: mixed sources (WRDS + Yahoo market fields)", p.ticker)
        else:
            logging.info("%s: Yahoo source used (WRDS unavailable/missing)", p.ticker)


def apply_overrides(peers: list[PeerRow], path: Path) -> None:
    if not path.exists():
        return
    df = pd.read_csv(path)
    if df.empty:
        return
    by_ticker = {p.ticker: p for p in peers}
    for _, row in df.iterrows():
        p = by_ticker.get(str(row.get("ticker", "")))
        if p is None:
            continue
        field_name = str(row.get("field", ""))
        year = row.get("year")
        value = row.get("value")
        source = str(row.get("source", "Override")).strip() or "Override"
        if pd.isna(value):
            continue
        if field_name in {"market_cap_ccy_m", "enterprise_value_ccy_m", "net_debt_ccy_m", "equity_beta", "gross_debt_ccy_m", "cash_ccy_m", "share_price_ccy", "fx_to_eur"}:
            setattr(p, field_name, float(value))
            p.source_by_field[field_name] = f"Override ({source})"
        elif field_name in {"revenue", "ebitda", "ebit"} and not pd.isna(year):
            getattr(p, field_name)[int(year)] = float(value)
            p.source_by_field[f"{field_name}_{int(year)}"] = f"Override ({source})"


def metric_multiple(ev: float | None, denom: float | None) -> float | None:
    if ev is None or denom in (None, 0):
        return None
    return ev / denom


def median(values: list[float]) -> float | None:
    clean = sorted(v for v in values if v is not None and not math.isnan(v))
    if not clean:
        return None
    n = len(clean)
    return clean[n // 2] if n % 2 else (clean[n // 2 - 1] + clean[n // 2]) / 2


def mean(values: list[float]) -> float | None:
    clean = [v for v in values if v is not None and not math.isnan(v)]
    return sum(clean) / len(clean) if clean else None


def compute_qc_rows(peers: list[PeerRow]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for p in peers:
        ev = p.enterprise_value_input_ccy_m
        ev_recon = None if p.market_cap_ccy_m is None or p.net_debt_ccy_m is None else p.market_cap_ccy_m + p.net_debt_ccy_m
        delta = None if ev is None or ev_recon is None else ev - ev_recon
        pct_delta = None if delta is None or ev in (None, 0) else delta / ev
        ev_ok = pct_delta is not None and abs(pct_delta) <= 0.05

        ev_ebitda_2024 = metric_multiple(ev, p.ebitda.get(2024))
        ev_ebit_2024 = metric_multiple(ev, p.ebit.get(2024))
        scale_flag = (ev_ebitda_2024 is not None and abs(ev_ebitda_2024) > 50) or (ev_ebit_2024 is not None and abs(ev_ebit_2024) > 80)

        missing = any(
            v is None
            for v in [p.market_cap_ccy_m, p.enterprise_value_ccy_m, p.net_debt_ccy_m, p.equity_beta]
            + [p.revenue[y] for y in FISCAL_YEARS]
            + [p.ebitda[y] for y in FISCAL_YEARS]
            + [p.ebit[y] for y in FISCAL_YEARS]
        )
        denom_issue = any(p.ebitda[y] in (None, 0) or p.ebit[y] in (None, 0) for y in FISCAL_YEARS)
        consistency_issue = (
            p.ebitda[2023] not in (None, 0)
            and p.ebitda[2024] is not None
            and (abs(p.ebitda[2024] / p.ebitda[2023]) > 10 or abs(p.ebitda[2024] / p.ebitda[2023]) < 0.1)
        )
        loss_year = any((p.ebit[y] or 0) < 0 for y in FISCAL_YEARS)

        out.append(
            {
                "company": p.company,
                "ticker": p.ticker,
                "delta_ev": delta,
                "pct_delta_ev": pct_delta,
                "ev_ebitda_2024": ev_ebitda_2024,
                "ev_ebit_2024": ev_ebit_2024,
                "checks": {
                    "ev_reconciliation": "PASS" if ev_ok else "FAIL",
                    "unit_scaling": "FAIL" if scale_flag else "PASS",
                    "missing_or_denominator": "FAIL" if missing or denom_issue else "PASS",
                    "year_consistency": "FAIL" if consistency_issue else "PASS",
                    "lossmaking": "FLAG" if loss_year else "PASS",
                },
            }
        )
    return out


def style_header(ws, row: int, start_col: int, end_col: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(start_col, end_col + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.font = font
        c.alignment = align


def build_workbook(peers: list[PeerRow], wrds_status: WrdsPullStatus) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Peer_Table"

    years = FISCAL_YEARS
    columns = [
        "Company", "Ticker", "Selected (1/0)", "Core Set (1/0)", "Peer Status", "Segment Fit", "Selection rationale", "Currency",
        "Share Price (CCY)", "Market Cap (CCY m)", "Enterprise Value (CCY m)", "Gross Debt (CCY m)", "Cash (CCY m)", "Net Debt (CCY m)", "Equity Beta", "FX to EUR",
    ]
    for y in years:
        columns.extend([f"Revenue {y} (CCY m)", f"EBITDA {y} (CCY m)", f"EBIT {y} (CCY m)", f"EV/Sales {y}", f"EV/EBITDA {y}", f"EV/EBIT {y}"])
    for i, c in enumerate(columns, 1):
        ws.cell(row=1, column=i, value=c)
    style_header(ws, 1, 1, len(columns))

    for r, p in enumerate(peers, 2):
        vals = [p.company, p.ticker, p.selected, p.core_set, p.peer_status, p.segment_fit, p.selection_rationale, p.currency,
                p.share_price_ccy, p.market_cap_ccy_m, p.enterprise_value_input_ccy_m, p.gross_debt_ccy_m, p.cash_ccy_m, p.net_debt_ccy_m, p.equity_beta, p.fx_to_eur]
        ev = p.enterprise_value_input_ccy_m
        for y in years:
            vals.extend([p.revenue[y], p.ebitda[y], p.ebit[y], metric_multiple(ev, p.revenue[y]), metric_multiple(ev, p.ebitda[y]), metric_multiple(ev, p.ebit[y])])
        for c, v in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=v)

    # Sources sheet with WRDS pull status + per-field source
    src = wb.create_sheet("Sources_and_AsOf")
    asof = AS_OF_OVERRIDE or datetime.now(timezone.utc).isoformat()
    src_meta = [
        ("As-of timestamp (UTC)", asof),
        ("WRDS connected (Y/N)", "Y" if wrds_status.connected else "N"),
        ("WRDS connection message", wrds_status.connection_message),
        ("WRDS mapping coverage (# peers with usable mapping)", wrds_status.mapping_coverage),
        ("WRDS statement filters", wrds_status.statement_filters),
        ("Mixed source fallback enabled", "Y" if ALLOW_MIXED_SOURCES else "N"),
        ("EV mode", "Provider EV" if USE_PROVIDER_EV_AS_TRUTH else "Computed EV = Market Cap + Net Debt"),
    ]
    for i, (k, v) in enumerate(src_meta, 1):
        src.cell(row=i, column=1, value=k)
        src.cell(row=i, column=2, value=v)

    peer_status_row = len(src_meta) + 2
    src.cell(row=peer_status_row, column=1, value="WRDS Pull Status by Peer")
    src.cell(row=peer_status_row, column=1).font = Font(bold=True)
    src.cell(row=peer_status_row + 1, column=1, value="Ticker")
    src.cell(row=peer_status_row + 1, column=2, value="Status")
    style_header(src, peer_status_row + 1, 1, 2)
    for i, p in enumerate(peers, peer_status_row + 2):
        src.cell(row=i, column=1, value=p.ticker)
        src.cell(row=i, column=2, value=wrds_status.per_peer_message.get(p.ticker, "No WRDS attempt"))

    field_header_row = peer_status_row + 3 + len(peers)
    headers = [
        "Company", "Ticker", "src_market_cap", "src_ev", "src_net_debt", "src_beta",
        "src_revenue_2023", "src_revenue_2024", "src_ebitda_2023", "src_ebitda_2024", "src_ebit_2023", "src_ebit_2024", "src_gross_debt", "src_cash", "src_currency", "src_fx",
    ]
    for i, h in enumerate(headers, 1):
        src.cell(row=field_header_row, column=i, value=h)
    style_header(src, field_header_row, 1, len(headers))

    def sf(p: PeerRow, key: str) -> str:
        return p.source_by_field.get(key, "MISSING SOURCE")

    for r, p in enumerate(peers, field_header_row + 1):
        vals = [
            p.company,
            p.ticker,
            sf(p, "market_cap_ccy_m"),
            sf(p, "enterprise_value_ccy_m"),
            sf(p, "net_debt_ccy_m"),
            sf(p, "equity_beta"),
            sf(p, "revenue_2023"),
            sf(p, "revenue_2024"),
            sf(p, "ebitda_2023"),
            sf(p, "ebitda_2024"),
            sf(p, "ebit_2023"),
            sf(p, "ebit_2024"),
            sf(p, "gross_debt_ccy_m"),
            sf(p, "cash_ccy_m"),
            sf(p, "currency"),
            sf(p, "fx_to_eur"),
        ]
        for c, v in enumerate(vals, 1):
            src.cell(row=r, column=c, value=v)

    # QC report
    qc = wb.create_sheet("QC_Report")
    qh = ["Company", "Ticker", "EV Reconciliation", "EV Delta", "EV Delta %", "Scaling", "Missing/Denominator", "Year Consistency", "Loss-making", "EV/EBITDA 2024", "EV/EBIT 2024"]
    for i, h in enumerate(qh, 1):
        qc.cell(row=1, column=i, value=h)
    style_header(qc, 1, 1, len(qh))
    for r, item in enumerate(compute_qc_rows(peers), 2):
        checks = item["checks"]
        vals = [item["company"], item["ticker"], checks["ev_reconciliation"], item["delta_ev"], item["pct_delta_ev"], checks["unit_scaling"], checks["missing_or_denominator"], checks["year_consistency"], checks["lossmaking"], item["ev_ebitda_2024"], item["ev_ebit_2024"]]
        for c, v in enumerate(vals, 1):
            qc.cell(row=r, column=c, value=v)

    # WACC
    wacc = wb.create_sheet("WACC_Model")
    wh = ["Company", "Selected", "Levered Beta", "Net Debt", "Market Cap", "D/E", "Unlevered Beta"]
    for i, h in enumerate(wh, 1):
        wacc.cell(row=1, column=i, value=h)
    style_header(wacc, 1, 1, len(wh))
    for r, p in enumerate(peers, 2):
        de = None if p.market_cap_ccy_m in (None, 0) or p.net_debt_ccy_m is None else p.net_debt_ccy_m / p.market_cap_ccy_m
        unlev = None if p.equity_beta is None or de is None else p.equity_beta / (1 + (1 - MARGINAL_TAX_RATE) * de)
        vals = [p.company, p.selected, p.equity_beta, p.net_debt_ccy_m, p.market_cap_ccy_m, de, unlev]
        for c, v in enumerate(vals, 1):
            wacc.cell(row=r, column=c, value=v)

    sel_unlev = [wacc.cell(row=i, column=7).value for i in range(2, 2 + len(peers)) if wacc.cell(row=i, column=2).value == 1 and wacc.cell(row=i, column=7).value is not None]
    sel_lev = [wacc.cell(row=i, column=3).value for i in range(2, 2 + len(peers)) if wacc.cell(row=i, column=2).value == 1 and wacc.cell(row=i, column=3).value is not None]
    relevered_beta = (median(sel_unlev) or 0) * (1 + (1 - MARGINAL_TAX_RATE) * TARGET_D_OVER_E)
    cost_equity = RISK_FREE_RATE + relevered_beta * EQUITY_RISK_PREMIUM + SMALL_FIRM_PREMIUM
    cost_debt_after_tax = COST_OF_DEBT_PRE_TAX * (1 - MARGINAL_TAX_RATE)
    debt_weight = TARGET_D_OVER_E / (1 + TARGET_D_OVER_E)
    equity_weight = 1 - debt_weight
    summary = [
        ("Beta methodology", f"{BETA_HORIZON} / {BETA_FREQUENCY} / {BETA_INDEX}"),
        ("Mean levered beta", mean(sel_lev)),
        ("Median levered beta", median(sel_lev)),
        ("Mean unlevered beta", mean(sel_unlev)),
        ("Median unlevered beta", median(sel_unlev)),
        ("Relevered beta", relevered_beta),
        ("Risk free", RISK_FREE_RATE),
        ("ERP", EQUITY_RISK_PREMIUM),
        ("Small firm premium", SMALL_FIRM_PREMIUM),
        ("Cost of equity", cost_equity),
        ("Cost of debt (pre-tax)", COST_OF_DEBT_PRE_TAX),
        ("Cost of debt methodology", COST_OF_DEBT_METHOD),
        ("WACC", equity_weight * cost_equity + debt_weight * cost_debt_after_tax),
        ("ERP source", ERP_SOURCE_NOTE),
        ("SFP source", SFP_SOURCE_NOTE),
    ]
    for i, (k, v) in enumerate(summary, len(peers) + 4):
        wacc.cell(row=i, column=1, value=k)
        wacc.cell(row=i, column=2, value=v)

    # Peer rationale and clean overview
    pr = wb.create_sheet("Peer_Rationale")
    ph = ["Company", "Ticker", "Segment Fit", "Role", "Selected", "Rationale"]
    for i, h in enumerate(ph, 1):
        pr.cell(row=1, column=i, value=h)
    style_header(pr, 1, 1, len(ph))
    for r, p in enumerate(peers, 2):
        vals = [p.company, p.ticker, p.segment_fit, p.peer_status, p.selected, p.selection_rationale]
        for c, v in enumerate(vals, 1):
            pr.cell(row=r, column=c, value=v)

    clean = wb.create_sheet("Clean_Overview")
    clean.cell(row=1, column=1, value="Weighted Average Cost of Capital").font = Font(bold=True)
    clean.cell(row=1, column=5, value="PEER GROUP (selected only)").font = Font(bold=True)
    lines = [
        ("Riskfree rate", RISK_FREE_RATE), ("Market risk premium", EQUITY_RISK_PREMIUM), ("Small firm premium", SMALL_FIRM_PREMIUM),
        ("Unlevered beta (median)", median(sel_unlev)), ("Relevered beta", relevered_beta), ("WACC", equity_weight * cost_equity + debt_weight * cost_debt_after_tax),
    ]
    for i, (k, v) in enumerate(lines, 3):
        clean.cell(row=i, column=1, value=k)
        clean.cell(row=i, column=2, value=v)
    headers = ["Company", "Levered Beta", "D/E", "Unlevered Beta"]
    for i, h in enumerate(headers, 5):
        clean.cell(row=3, column=i, value=h)
    style_header(clean, 3, 5, 8)
    row = 4
    for p in peers:
        if p.selected != 1:
            continue
        de = None if p.market_cap_ccy_m in (None, 0) or p.net_debt_ccy_m is None else p.net_debt_ccy_m / p.market_cap_ccy_m
        ub = None if p.equity_beta is None or de is None else p.equity_beta / (1 + (1 - MARGINAL_TAX_RATE) * de)
        clean.cell(row=row, column=5, value=p.company)
        clean.cell(row=row, column=6, value=p.equity_beta)
        clean.cell(row=row, column=7, value=de)
        clean.cell(row=row, column=8, value=ub)
        row += 1

    for sh in wb.worksheets:
        sh.freeze_panes = sh.freeze_panes or "A2"
    wb.save(OUTPUT_FILE)
    logging.info("Saved workbook to %s", OUTPUT_FILE)


def main() -> None:
    setup_logging()
    peers = parse_peers(PEER_INPUT_FILE)
    wrds_mapping = parse_wrds_mapping(WRDS_MAPPING_FILE)
    wrds_status = fetch_from_wrds(peers, wrds_mapping)
    logging.info("WRDS status: connected=%s, mapping_coverage=%s, message=%s", wrds_status.connected, wrds_status.mapping_coverage, wrds_status.connection_message)
    fetch_from_yahoo(peers, wrds_status)
    apply_overrides(peers, OVERRIDE_FILE)
    build_workbook(peers, wrds_status)


if __name__ == "__main__":
    main()
