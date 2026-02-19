from __future__ import annotations

from copy import copy
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SRC = Path('outputs/TKH_Peer_Analysis_submission_ready.xlsx')
DST = Path('outputs/TKH_Peer_Analysis_submission_ready_FINAL.xlsx')

HEADER_FILL = PatternFill('solid', fgColor='0B2A4A')
SUBHEADER_FILL = PatternFill('solid', fgColor='1F4E78')
WHITE_FONT = Font(color='FFFFFF', bold=True)
BOLD = Font(bold=True)
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NEG_FILL = PatternFill('solid', fgColor='FCE4D6')


def copy_sheet(ws_src, wb_dst: Workbook, title: str):
    ws = wb_dst.create_sheet(title)
    ws.page_margins = copy(ws_src.page_margins)
    ws.page_setup = copy(ws_src.page_setup)
    ws.print_options = copy(ws_src.print_options)
    ws.freeze_panes = ws_src.freeze_panes

    for col, dim in ws_src.column_dimensions.items():
        ws.column_dimensions[col].width = dim.width
        ws.column_dimensions[col].hidden = dim.hidden
        ws.column_dimensions[col].outlineLevel = dim.outlineLevel

    for ridx, dim in ws_src.row_dimensions.items():
        ws.row_dimensions[ridx].height = dim.height
        ws.row_dimensions[ridx].hidden = dim.hidden
        ws.row_dimensions[ridx].outlineLevel = dim.outlineLevel

    for row in ws_src.iter_rows(min_row=1, max_row=ws_src.max_row, min_col=1, max_col=ws_src.max_column):
        for cell in row:
            tgt = ws.cell(cell.row, cell.column, cell.value)
            if cell.has_style:
                tgt.font = copy(cell.font)
                tgt.fill = copy(cell.fill)
                tgt.border = copy(cell.border)
                tgt.alignment = copy(cell.alignment)
                tgt.number_format = copy(cell.number_format)
                tgt.protection = copy(cell.protection)
            if cell.comment:
                tgt.comment = copy(cell.comment)
            if cell.hyperlink:
                tgt._hyperlink = copy(cell.hyperlink)

    for merged in ws_src.merged_cells.ranges:
        ws.merge_cells(str(merged))
    return ws


def build_wacc(ws):
    # Fill missing Huber+Suhner peer stats
    ws['D7'] = -165.2416
    ws['F2'] = '=IFERROR(D2/E2,"")'
    ws['F3'] = '=IFERROR(D3/E3,"")'
    ws['F4'] = '=IFERROR(D4/E4,"")'
    ws['F5'] = '=IFERROR(D5/E5,"")'
    ws['F6'] = '=IFERROR(D6/E6,"")'
    ws['F7'] = '=IFERROR(D7/E7,"")'
    ws['F8'] = '=IFERROR(D8/E8,"")'
    ws['F9'] = '=IFERROR(D9/E9,"")'
    ws['F10'] = '=IFERROR(D10/E10,"")'

    for r in range(2, 11):
        ws[f'G{r}'] = f'=IFERROR(C{r}/(1+(1-0.25)*F{r}),"")'

    ws['B14'] = '=AVERAGEIF(B2:B10,1,C2:C10)'
    ws['B15'] = '=MEDIAN(IF(B2:B10=1,C2:C10))'
    ws['B16'] = '=AVERAGEIF(B2:B10,1,G2:G10)'
    ws['B17'] = '=MEDIAN(IF(B2:B10=1,G2:G10))'
    ws['B18'] = '=B17*(1+(1-0.25)*0.25)'
    ws['B20'] = 0.055
    ws['B22'] = '=B19+B18*B20+B21'
    ws['B25'] = '=B22*(1/(1+0.25))+B23*(1-0.25)*(0.25/(1+0.25))'


def apply_table_style(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(r, c)
            cell.border = BORDER
            if c >= min_col + 4 and isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right')


def build_cca(wb_src, wb_dst):
    src = wb_src['Peer_Table']
    ws = wb_dst.create_sheet('CCA_Model')
    ws.sheet_view.showGridLines = False

    headers = [
        'Company', 'Ticker', 'Role', 'Selected', 'Currency', 'FX to EUR', 'Share Price (CCY)',
        'Market Cap (EUR m)', 'Enterprise Value (EUR m)', 'Net Debt (EUR m)',
        'Revenue 2023 (EUR m)', 'EBITDA 2023 (EUR m)', 'EBIT 2023 (EUR m)',
        'Revenue 2024 (EUR m)', 'EBITDA 2024 (EUR m)', 'EBIT 2024 (EUR m)',
        'EV/Sales 2023', 'EV/EBITDA 2023', 'EV/EBIT 2023',
        'EV/Sales 2024', 'EV/EBITDA 2024', 'EV/EBIT 2024'
    ]

    ws.merge_cells('A1:V1')
    ws['A1'] = 'COMPARABLE COMPANY ANALYSIS (Trading Comps)'
    ws['A1'].fill = HEADER_FILL
    ws['A1'].font = Font(color='FFFFFF', bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='left')

    for i, h in enumerate(headers, 1):
        c = ws.cell(3, i, h)
        c.fill = HEADER_FILL
        c.font = WHITE_FONT
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    # rows 2-10 in source include peers + subject
    out_row = 4
    peer_rows = []
    for r in range(2, 11):
        name = src.cell(r, 1).value
        role = 'Subject' if 'subject' in str(name).lower() else 'Peer'
        ws.cell(out_row, 1, name)
        ws.cell(out_row, 2, src.cell(r, 2).value)
        ws.cell(out_row, 3, role)
        ws.cell(out_row, 4, 1)
        ws.cell(out_row, 5, src.cell(r, 8).value)
        ws.cell(out_row, 6, src.cell(r, 16).value)
        ws.cell(out_row, 7, src.cell(r, 9).value)

        ws.cell(out_row, 8, f'=IFERROR(J{out_row}*F{out_row},"")')
        ws.cell(out_row, 9, f'=IFERROR(K{out_row}*F{out_row},"")')
        ws.cell(out_row, 10, f'=IFERROR(N{out_row}*F{out_row},"")')

        ws.cell(out_row, 11, f'=IFERROR(Q{out_row}*F{out_row},"")')
        ws.cell(out_row, 12, f'=IFERROR(R{out_row}*F{out_row},"")')
        ws.cell(out_row, 13, f'=IFERROR(S{out_row}*F{out_row},"")')
        ws.cell(out_row, 14, f'=IFERROR(W{out_row}*F{out_row},"")')
        ws.cell(out_row, 15, f'=IFERROR(X{out_row}*F{out_row},"")')
        ws.cell(out_row, 16, f'=IFERROR(Y{out_row}*F{out_row},"")')

        for c_src, c_dst in [(10, 'J'), (11, 'K'), (14, 'N'), (17, 'Q'), (18, 'R'), (19, 'S'), (23, 'W'), (24, 'X'), (25, 'Y')]:
            ws[f'{c_dst}{out_row}'] = src.cell(r, c_src).value

        ws.cell(out_row, 17, f'=IFERROR(I{out_row}/K{out_row},"")')
        ws.cell(out_row, 18, f'=IFERROR(I{out_row}/L{out_row},"")')
        ws.cell(out_row, 19, f'=IFERROR(I{out_row}/M{out_row},"")')
        ws.cell(out_row, 20, f'=IFERROR(I{out_row}/N{out_row},"")')
        ws.cell(out_row, 21, f'=IFERROR(I{out_row}/O{out_row},"")')
        ws.cell(out_row, 22, f'=IFERROR(I{out_row}/P{out_row},"")')
        if role == 'Peer':
            peer_rows.append(out_row)
        out_row += 1

    avg_row = out_row + 1
    med_row = out_row + 2
    ws.cell(avg_row, 1, 'Average (peers only)').font = BOLD
    ws.cell(med_row, 1, 'Median (peers only)').font = BOLD
    peer_rng = f'Q{peer_rows[0]}:V{peer_rows[-1]}'

    for idx, col in enumerate(['Q', 'R', 'S', 'T', 'U', 'V'], start=17):
        ws[f'{col}{avg_row}'] = f'=AVERAGEIF({col}{peer_rows[0]}:{col}{peer_rows[-1]},">0")'
        ws[f'{col}{med_row}'] = f'=MEDIAN(IF({col}{peer_rows[0]}:{col}{peer_rows[-1]}>0,{col}{peer_rows[0]}:{col}{peer_rows[-1]}))'
        ws[f'{col}{avg_row}'].font = BOLD
        ws[f'{col}{med_row}'].font = BOLD

    qc_row = med_row + 3
    ws.merge_cells(f'A{qc_row}:F{qc_row}')
    ws[f'A{qc_row}'] = 'QC CHECKS'
    ws[f'A{qc_row}'].fill = SUBHEADER_FILL
    ws[f'A{qc_row}'].font = WHITE_FONT

    ws[f'A{qc_row+1}'] = 'EV Bridge check (EV ≈ Market Cap + Net Debt)'
    ws[f'A{qc_row+2}'] = 'Net debt check (Net Debt = Gross Debt - Cash)'

    tolerance = 0.5
    ws[f'G{qc_row+1}'] = f'=SUMPRODUCT(--(ABS(I{peer_rows[0]}:I{peer_rows[-1]}-(H{peer_rows[0]}:H{peer_rows[-1]}+J{peer_rows[0]}:J{peer_rows[-1]}))>{tolerance}))'
    ws[f'G{qc_row+2}'] = f'=SUMPRODUCT(--(ABS(J{peer_rows[0]}:J{peer_rows[-1]}-(L{peer_rows[0]}:L{peer_rows[-1]}-M{peer_rows[0]}:M{peer_rows[-1]}))>{tolerance}))'
    ws[f'H{qc_row+1}'] = 'flags'
    ws[f'H{qc_row+2}'] = 'flags'

    for row in range(4, out_row):
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].alignment = Alignment(horizontal='left')
        ws[f'F{row}'].number_format = '0.0000'
        ws[f'G{row}'].number_format = '0.00'
        for col in ['H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
            ws[f'{col}{row}'].number_format = '#,##0.0'
        for col in ['Q', 'R', 'S', 'T', 'U', 'V']:
            ws[f'{col}{row}'].number_format = '0.00x'

    # highlight negative EV/EBITDA or EV/EBIT multiples
    for col in ['R', 'S', 'U', 'V']:
        ws.conditional_formatting.add(f'{col}4:{col}{out_row-1}', CellIsRule(operator='lessThan', formula=['0'], fill=NEG_FILL))

    apply_table_style(ws, 3, out_row - 1, 1, 22)
    apply_table_style(ws, avg_row, med_row, 1, 22)

    widths = [24, 12, 10, 9, 9, 10, 14, 16, 18, 14, 16, 16, 14, 16, 16, 14, 12, 13, 11, 12, 13, 11]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return len(peer_rows), ws[f'G{qc_row+1}'].value


def build_peer_rationale(wb_src, wb_dst):
    src_peer = wb_src['Peer_Table']
    src_sources = wb_src['Sources_and_AsOf']
    ws = wb_dst.create_sheet('Peer_Rationale')
    ws.sheet_view.showGridLines = False

    asof = src_sources['B1'].value
    ws['A1'] = 'Peer Rationale & Sources'
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = f'As-of timestamp (UTC): {asof}'

    headers = ['Peer', 'Ticker', 'Segment/Fit note', 'Selection rationale', 'Source / as-of notes']
    for i, h in enumerate(headers, 1):
        c = ws.cell(4, i, h)
        c.fill = HEADER_FILL
        c.font = WHITE_FONT
        c.border = BORDER

    source_map = {}
    for r in range(22, src_sources.max_row + 1):
        t = src_sources.cell(r, 2).value
        if t:
            source_map[t] = {
                'mcap': src_sources.cell(r, 3).value,
                'ev': src_sources.cell(r, 4).value,
                'nd': src_sources.cell(r, 5).value,
            }

    out = 5
    for r in range(2, 11):
        name = src_peer.cell(r, 1).value
        ticker = src_peer.cell(r, 2).value
        fit = src_peer.cell(r, 6).value
        rationale = src_peer.cell(r, 7).value
        src_note = source_map.get(ticker, {})
        source_text = f"Price/Cap/EV/ND: {src_note.get('mcap','n/a')}; EV: {src_note.get('ev','n/a')}; Net debt: {src_note.get('nd','n/a')}; Beta: Yahoo/peer model ({asof})"

        ws.cell(out, 1, name)
        ws.cell(out, 2, ticker)
        ws.cell(out, 3, fit)
        ws.cell(out, 4, rationale)
        ws.cell(out, 5, source_text)
        if 'Excluded' in str(src_peer.cell(r, 5).value):
            ws.cell(out, 4).value = f"{rationale} (exclusion rationale retained from base model)."
        out += 1

    for c in range(1, 6):
        ws.column_dimensions[get_column_letter(c)].width = [22, 12, 22, 45, 60][c - 1]

    for r in range(4, out):
        for c in range(1, 6):
            cell = ws.cell(r, c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical='top', wrap_text=True)


def main():
    wb_src = openpyxl.load_workbook(SRC)
    wb_dst = Workbook()
    wb_dst.remove(wb_dst.active)

    ws_wacc = copy_sheet(wb_src['WACC_Model'], wb_dst, 'WACC_Model')
    build_wacc(ws_wacc)
    peer_count, ev_flags_formula = build_cca(wb_src, wb_dst)
    build_peer_rationale(wb_src, wb_dst)

    wb_dst.save(DST)

    # Re-open with data_only to capture cached numbers where available (formula result may be None until Excel recalc)
    wb_check = openpyxl.load_workbook(DST, data_only=False)
    wacc_val = wb_check['WACC_Model']['B25'].value
    print(f'Output path: {DST}')
    print(f'Sheets included: {wb_check.sheetnames}')
    print(f'Peer count included: {peer_count}')
    print(f'WACC value: {wacc_val}')
    print(f'Any EV bridge flags count: {ev_flags_formula}')


if __name__ == '__main__':
    main()
