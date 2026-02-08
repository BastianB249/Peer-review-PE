from __future__ import annotations

from typing import Any, Dict, Iterable

import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

INPUT_FILE = "TKH_Peer_Analysis.xlsx"
OUTPUT_FILE = "TKH_Peer_Analysis_filled.xlsx"
SHEET_NAME = "Peer_Table"

HEADER_ROW_1 = 1
HEADER_ROW_2 = 2
DATA_START_ROW = 3

MISSING_OPERATING_FILL = PatternFill("solid", fgColor="FFF2CC")

# Map “your” tickers -> Yahoo tickers when needed
TICKER_MAP = {
    "COGX": "CGNX",        # Cognex
    "ASMI.AS": "ASM.AS",   # ASM International
    # Add TKH mapping only if your sheet uses "TKH" as ticker.
    # If your sheet uses the correct Yahoo symbol already (e.g. TWEKA.AS), you don't need this.
    "TKH": "TWEKA.AS",     # TKH Group
}

REVENUE_LABELS = ["Total Revenue", "TotalRevenue"]
EBIT_LABELS = ["EBIT", "Operating Income", "OperatingIncome"]
EBITDA_LABELS = ["EBITDA"]
DA_LABELS = [
    "Depreciation & Amortization",
    "Depreciation And Amortization",
    "Depreciation/Amortization",
]
TOTAL_DEBT_LABELS = [
    "Total Debt",
]
DEBT_COMPONENT_LABELS = [
    "Long Term Debt",
    "Short Long Term Debt",
    "Short Term Debt",
    "Current Debt",
]
CASH_LABELS = [
    "Cash And Cash Equivalents",
    "Cash",
    "Cash And Cash Equivalents Including Short Term Investments",
]


def _map_ticker(ticker: str) -> str:
    return TICKER_MAP.get(ticker.strip(), ticker.strip())


def _to_ccy_m(value: Any) -> float | None:
    """Convert absolute currency units -> currency millions."""
    if value in (None, ""):
        return None
    try:
        return float(value) / 1_000_000
    except (TypeError, ValueError):
        return None


def _last_close_price(tkr: yf.Ticker) -> float | None:
    """Get last close price from recent trading days (more robust than info)."""
    try:
        hist = tkr.history(period="5d")
        if hist is None or hist.empty:
            return None
        close = hist["Close"].dropna()
        if close.empty:
            return None
        return float(close.iloc[-1])
    except Exception:
        return None


def _find_row_label(index: Iterable[Any], labels: Iterable[str]) -> Any | None:
    """Find best-matching row label in a yfinance financials dataframe index."""
    label_map = {str(label).lower(): label for label in index}
    for candidate in labels:
        key = candidate.lower()
        if key in label_map:
            return label_map[key]
    return None


def _extract_metric_by_year(financials: pd.DataFrame, labels: Iterable[str]) -> Dict[int, float]:
    """Return {year: value} for a given metric row from tkr.financials."""
    if financials is None or financials.empty:
        return {}
    row_label = _find_row_label(financials.index, labels)
    if row_label is None:
        return {}
    series = financials.loc[row_label]

    data: Dict[int, float] = {}
    for col, value in series.items():
        if pd.isna(value):
            continue
        try:
            year = pd.Timestamp(col).year
        except Exception:
            try:
                year = int(str(col)[:4])
            except Exception:
                continue
        try:
            data[int(year)] = float(value)
        except Exception:
            continue
    return data


def _latest_balance_sheet_column(balance_sheet: pd.DataFrame) -> Any | None:
    if balance_sheet is None or balance_sheet.empty:
        return None
    columns = list(balance_sheet.columns)
    if not columns:
        return None
    try:
        return max(columns)
    except Exception:
        return columns[0]


def _extract_balance_value(balance_sheet: pd.DataFrame, labels: Iterable[str]) -> float | None:
    if balance_sheet is None or balance_sheet.empty:
        return None
    row_label = _find_row_label(balance_sheet.index, labels)
    if row_label is None:
        return None
    latest_col = _latest_balance_sheet_column(balance_sheet)
    if latest_col is None:
        return None
    value = balance_sheet.loc[row_label].get(latest_col)
    if value is None or pd.isna(value):
        return None
    try:
        return float(value)
    except Exception:
        return None


def _compute_net_debt(info: Dict[str, Any], balance_sheet: pd.DataFrame) -> float | None:
    net_debt = info.get("netDebt")
    if net_debt not in (None, ""):
        try:
            return float(net_debt)
        except Exception:
            pass

    total_debt = _extract_balance_value(balance_sheet, TOTAL_DEBT_LABELS)
    if total_debt is None:
        total_debt = 0.0
        found = False
        for label in DEBT_COMPONENT_LABELS:
            component = _extract_balance_value(balance_sheet, [label])
            if component is not None:
                total_debt += component
                found = True
        if not found:
            total_debt = None

    cash_value = _extract_balance_value(balance_sheet, CASH_LABELS)
    if total_debt is None or cash_value is None:
        return None
    return total_debt - cash_value


def _extract_years(group_map: Dict[tuple[str, str], int], label: str) -> list[int]:
    """Read available years from the 2-row header group map (e.g. Revenue (CCY m) / 2023, 2024)."""
    years: list[int] = []
    for (group_label, year_label) in group_map:
        if group_label == label:
            try:
                years.append(int(year_label))
            except (TypeError, ValueError):
                continue
    return sorted(years)


def _build_header_maps(ws) -> tuple[Dict[str, int], Dict[tuple[str, str], int]]:
    """
    Build:
      base_cols: { "Ticker": col, ... } for columns whose row2 header is blank
      group_cols: { ("Revenue (CCY m)", "2023"): col, ... } for grouped year columns

    Handles merged cells in row1 where only the leftmost cell contains the group label.
    """
    base_cols: Dict[str, int] = {}
    group_cols: Dict[tuple[str, str], int] = {}

    last_header_1: str | None = None

    for col in range(1, ws.max_column + 1):
        header_1 = ws.cell(row=HEADER_ROW_1, column=col).value
        header_2 = ws.cell(row=HEADER_ROW_2, column=col).value

        if header_1 is not None and str(header_1).strip() != "":
            last_header_1 = str(header_1).strip()

        if last_header_1 is None:
            continue

        if header_2 is None or str(header_2).strip() == "":
            base_cols[last_header_1] = col
        else:
            group_cols[(last_header_1, str(header_2).strip())] = col

    return base_cols, group_cols


def _write_operating_value(ws, row: int, col: int, value: float | None) -> None:
    cell = ws.cell(row=row, column=col)
    if value is None:
        cell.value = None
        cell.fill = MISSING_OPERATING_FILL
    else:
        cell.value = _to_ccy_m(value)


def _fetch_ticker_data(ysym: str) -> tuple[yf.Ticker | None, dict, pd.DataFrame, pd.DataFrame]:
    try:
        tkr = yf.Ticker(ysym)
    except Exception as exc:
        print(f"{ysym}: ticker init failed: {exc}")
        return None, {}, pd.DataFrame(), pd.DataFrame()

    try:
        info = tkr.get_info() or {}
    except Exception as exc:
        print(f"{ysym}: get_info failed: {exc}")
        info = {}

    try:
        financials = tkr.financials
    except Exception as exc:
        print(f"{ysym}: financials failed: {exc}")
        financials = pd.DataFrame()

    try:
        balance_sheet = tkr.balance_sheet
    except Exception as exc:
        print(f"{ysym}: balance sheet failed: {exc}")
        balance_sheet = pd.DataFrame()

    return tkr, info, financials, balance_sheet


def _fetch_fx_rate(ccy: str | None, fx_cache: Dict[str, float]) -> float | None:
    if not ccy:
        return None
    if ccy == "EUR":
        return 1.0
    if ccy in fx_cache:
        return fx_cache[ccy]

    direct_symbol = f"{ccy}EUR=X"
    direct_ticker = yf.Ticker(direct_symbol)
    direct_rate = _last_close_price(direct_ticker)
    if direct_rate is not None:
        fx_cache[ccy] = float(direct_rate)
        return fx_cache[ccy]

    inverse_symbol = f"EUR{ccy}=X"
    inverse_ticker = yf.Ticker(inverse_symbol)
    inverse_rate = _last_close_price(inverse_ticker)
    if inverse_rate is not None:
        fx_cache[ccy] = 1.0 / float(inverse_rate)
        return fx_cache[ccy]

    return None


def _find_tkh_inputs_block(ws) -> int | None:
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value and str(cell_value).strip() == "TKH Inputs":
            return row
    return None


def _parse_year_columns(ws, header_row: int) -> Dict[int, int]:
    year_cols: Dict[int, int] = {}
    for col in range(2, ws.max_column + 1):
        header_value = ws.cell(row=header_row, column=col).value
        if header_value is None or str(header_value).strip() == "":
            continue
        try:
            year = int(str(header_value).strip())
        except ValueError:
            continue
        year_cols[year] = col
    return year_cols


def _fill_tkh_inputs(
    ws,
    info: Dict[str, Any],
    financials: pd.DataFrame,
    balance_sheet: pd.DataFrame,
) -> None:
    block_row = _find_tkh_inputs_block(ws)
    if block_row is None:
        print("TKH Inputs block not found; skipping.")
        return

    header_row = block_row + 1
    first_metric_row = header_row + 1
    year_cols = _parse_year_columns(ws, header_row)
    if not year_cols:
        print("TKH Inputs year headers not found; skipping.")
        return

    latest_year = max(year_cols)
    prior_year = min(year_cols)

    metric_rows: Dict[str, int] = {}
    for row in range(first_metric_row, ws.max_row + 1):
        label = ws.cell(row=row, column=1).value
        if label is None or str(label).strip() == "":
            break
        metric_rows[str(label).strip()] = row

    revenue_by_year = _extract_metric_by_year(financials, REVENUE_LABELS)
    ebitda_by_year = _extract_metric_by_year(financials, EBITDA_LABELS)
    ebit_by_year = _extract_metric_by_year(financials, EBIT_LABELS)

    if not ebitda_by_year:
        da_by_year = _extract_metric_by_year(financials, DA_LABELS)
        for year in set(ebit_by_year) & set(da_by_year):
            ebitda_by_year[year] = ebit_by_year[year] + da_by_year[year]

    metric_sources = {
        "Revenue (CCY m)": revenue_by_year,
        "EBITDA (CCY m)": ebitda_by_year,
        "EBIT (CCY m)": ebit_by_year,
    }

    for metric_label, data in metric_sources.items():
        row = metric_rows.get(metric_label)
        if row is None:
            continue
        for year, col in year_cols.items():
            value = data.get(year)
            ws.cell(row=row, column=col, value=_to_ccy_m(value))

    net_debt_row = metric_rows.get("Net Debt (CCY m)")
    if net_debt_row is not None:
        net_debt_value = _compute_net_debt(info, balance_sheet)
        ws.cell(row=net_debt_row, column=year_cols[latest_year], value=_to_ccy_m(net_debt_value))

    shares_row = metric_rows.get("Shares Outstanding (m)")
    if shares_row is not None:
        shares_value = info.get("sharesOutstanding") or info.get("floatShares")
        shares_m = shares_value / 1_000_000 if shares_value else None
        ws.cell(row=shares_row, column=year_cols[latest_year], value=shares_m)
        if prior_year != latest_year:
            ws.cell(row=shares_row, column=year_cols[prior_year], value=shares_m)

    adjustments_row = metric_rows.get("Adjustments (CCY m)")
    if adjustments_row is not None:
        for col in year_cols.values():
            ws.cell(row=adjustments_row, column=col, value=0)


def main() -> None:
    wb = load_workbook(INPUT_FILE)
    ws = wb[SHEET_NAME]

    base_cols, group_cols = _build_header_maps(ws)

    required_base = [
        "Ticker",
        "Currency",
        "Share Price (CCY)",
        "Market Cap (CCY m)",
        "Enterprise Value (CCY m)",
        "Net Debt (CCY m)",
        "FX to EUR",
        "Share Price (EUR)",
        "Market Cap (EUR m)",
        "Enterprise Value (EUR m)",
        "Net Debt (EUR m)",
    ]
    missing = [c for c in required_base if c not in base_cols]
    if missing:
        raise ValueError(f"Missing required base columns (row {HEADER_ROW_1}): {missing}")

    years = _extract_years(group_cols, "Revenue (CCY m)")
    if len(years) < 2:
        raise ValueError("Expected two fiscal years in the header for Revenue (CCY m).")

    for y in years:
        for group in ("Revenue (CCY m)", "EBITDA (CCY m)", "EBIT (CCY m)"):
            if (group, str(y)) not in group_cols:
                raise ValueError(f"Missing grouped column: {group} / {y}")

    ticker_cache: Dict[str, tuple[yf.Ticker | None, Dict[str, Any], pd.DataFrame, pd.DataFrame]] = {}
    fx_cache: Dict[str, float] = {}

    for row in range(DATA_START_ROW, ws.max_row + 1):
        ticker_val = ws.cell(row=row, column=base_cols["Ticker"]).value
        if not ticker_val:
            continue

        raw = str(ticker_val).strip()
        ysym = _map_ticker(raw)
        if ysym not in ticker_cache:
            ticker_cache[ysym] = _fetch_ticker_data(ysym)

        tkr, info, financials, balance_sheet = ticker_cache[ysym]
        if tkr is None:
            continue

        share_price = _last_close_price(tkr)
        currency = info.get("currency")
        market_cap = info.get("marketCap")
        enterprise_value = info.get("enterpriseValue")
        net_debt = _compute_net_debt(info, balance_sheet)
        fx_rate = _fetch_fx_rate(currency, fx_cache)

        revenue_by_year = _extract_metric_by_year(financials, REVENUE_LABELS)
        ebit_by_year = _extract_metric_by_year(financials, EBIT_LABELS)
        ebitda_by_year = _extract_metric_by_year(financials, EBITDA_LABELS)

        # If EBITDA missing, try EBIT + D&A
        if not ebitda_by_year:
            da_by_year = _extract_metric_by_year(financials, DA_LABELS)
            for year in set(ebit_by_year) & set(da_by_year):
                ebitda_by_year[year] = ebit_by_year[year] + da_by_year[year]

        if currency is None:
            print(f"{raw} -> {ysym}: warning - missing currency")
        if market_cap is None:
            print(f"{raw} -> {ysym}: warning - missing market cap")
        if enterprise_value is None:
            print(f"{raw} -> {ysym}: warning - missing enterprise value")
        if net_debt is None:
            print(f"{raw} -> {ysym}: warning - missing net debt")
        if fx_rate is None:
            print(f"{raw} -> {ysym}: warning - missing FX rate to EUR")
        if not revenue_by_year:
            print(f"{raw} -> {ysym}: warning - missing revenue in financials")
        if not ebitda_by_year:
            print(f"{raw} -> {ysym}: warning - missing EBITDA in financials")
        if not ebit_by_year:
            print(f"{raw} -> {ysym}: warning - missing EBIT in financials")

        # Write base cells
        ws.cell(row=row, column=base_cols["Currency"], value=currency)
        ws.cell(row=row, column=base_cols["Share Price (CCY)"], value=share_price)
        ws.cell(row=row, column=base_cols["Market Cap (CCY m)"], value=_to_ccy_m(market_cap))
        ws.cell(row=row, column=base_cols["Enterprise Value (CCY m)"], value=_to_ccy_m(enterprise_value))
        ws.cell(row=row, column=base_cols["Net Debt (CCY m)"], value=_to_ccy_m(net_debt))
        ws.cell(row=row, column=base_cols["FX to EUR"], value=fx_rate)
        ws.cell(
            row=row,
            column=base_cols["Share Price (EUR)"],
            value=share_price * fx_rate if share_price is not None and fx_rate is not None else None,
        )
        ws.cell(
            row=row,
            column=base_cols["Market Cap (EUR m)"],
            value=_to_ccy_m(market_cap) * fx_rate
            if market_cap is not None and fx_rate is not None
            else None,
        )
        ws.cell(
            row=row,
            column=base_cols["Enterprise Value (EUR m)"],
            value=_to_ccy_m(enterprise_value) * fx_rate
            if enterprise_value is not None and fx_rate is not None
            else None,
        )
        ws.cell(
            row=row,
            column=base_cols["Net Debt (EUR m)"],
            value=_to_ccy_m(net_debt) * fx_rate if net_debt is not None and fx_rate is not None else None,
        )

        # Write year-group operating metrics
        for year in years:
            _write_operating_value(ws, row, group_cols[("Revenue (CCY m)", str(year))], revenue_by_year.get(year))
            _write_operating_value(ws, row, group_cols[("EBITDA (CCY m)", str(year))], ebitda_by_year.get(year))
            _write_operating_value(ws, row, group_cols[("EBIT (CCY m)", str(year))], ebit_by_year.get(year))

        print(f"Filled {raw} (Yahoo: {ysym})")

    if "TWEKA.AS" not in ticker_cache:
        ticker_cache["TWEKA.AS"] = _fetch_ticker_data("TWEKA.AS")
    _, tkh_info, tkh_financials, tkh_balance = ticker_cache["TWEKA.AS"]
    _fill_tkh_inputs(ws, tkh_info, tkh_financials, tkh_balance)

    wb.save(OUTPUT_FILE)
    print(f"Saved {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
